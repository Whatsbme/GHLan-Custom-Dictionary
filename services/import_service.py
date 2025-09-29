"""
导入服务
Import Service
"""

import os
import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging

from openpyxl import load_workbook
from services.dictionary_service import dictionary_service


class ImportService:
    """导入服务类"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def import_from_excel(self, filepath: str, progress_callback=None) -> Dict[str, Any]:
        """从Excel文件导入词条"""
        try:
            results = {
                'success': 0,
                'failed': 0,
                'errors': [],
                'total': 0
            }
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # 读取标题行
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # 计算总行数
            total_rows = ws.max_row - 1  # 减去标题行
            results['total'] = total_rows
            
            # 读取数据行
            current_row = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                current_row += 1
                
                # 更新进度
                if progress_callback:
                    word_id = row[0] if row and row[0] else f"第{current_row}行"
                    progress_callback(current_row, total_rows, str(word_id))
                
                try:
                    word_data = self._parse_excel_row(headers, row)
                    if word_data:
                        success = dictionary_service.add_word_entry(word_data)
                        if success:
                            results['success'] += 1
                        else:
                            results['failed'] += 1
                            results['errors'].append(f"添加词条失败: {word_data.get('word_id', 'Unknown')}")
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"解析行数据失败: {str(e)}")
            
            self.logger.info(f"Excel导入完成: 成功{results['success']}条, 失败{results['failed']}条")
            return results
            
        except Exception as e:
            self.logger.error(f"Excel导入失败: {e}")
            raise
    
    def import_from_csv(self, filepath: str, progress_callback=None) -> Dict[str, Any]:
        """从CSV文件导入词条"""
        try:
            results = {
                'success': 0,
                'failed': 0,
                'errors': [],
                'total': 0
            }
            
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)  # 读取所有行
                total_rows = len(rows)
                results['total'] = total_rows
                
                for index, row in enumerate(rows):
                    current_row = index + 1
                    
                    # 更新进度
                    if progress_callback:
                        word_id = row.get('word_id', f"第{current_row}行")
                        progress_callback(current_row, total_rows, str(word_id))
                    
                    try:
                        word_data = self._parse_csv_row(row)
                        if word_data:
                            success = dictionary_service.add_word_entry(word_data)
                            if success:
                                results['success'] += 1
                            else:
                                results['failed'] += 1
                                results['errors'].append(f"添加词条失败: {word_data.get('word_id', 'Unknown')}")
                    except Exception as e:
                        results['failed'] += 1
                        results['errors'].append(f"解析行数据失败: {str(e)}")
            
            self.logger.info(f"CSV导入完成: 成功{results['success']}条, 失败{results['failed']}条")
            return results
            
        except Exception as e:
            self.logger.error(f"CSV导入失败: {e}")
            raise
    
    def import_from_json(self, filepath: str) -> Dict[str, Any]:
        """从JSON文件导入词条"""
        try:
            results = {
                'success': 0,
                'failed': 0,
                'errors': []
            }
            
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 支持单个词条或词条列表
            if isinstance(data, dict):
                data = [data]
            elif not isinstance(data, list):
                raise ValueError("JSON文件格式不正确")
            
            for word_data in data:
                try:
                    if self._validate_word_data(word_data):
                        success = dictionary_service.add_word_entry(word_data)
                        if success:
                            results['success'] += 1
                        else:
                            results['failed'] += 1
                            results['errors'].append(f"添加词条失败: {word_data.get('word_id', 'Unknown')}")
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"处理词条失败: {str(e)}")
            
            self.logger.info(f"JSON导入完成: 成功{results['success']}条, 失败{results['failed']}条")
            return results
            
        except Exception as e:
            self.logger.error(f"JSON导入失败: {e}")
            raise
    
    def _parse_excel_row(self, headers: List[str], row: tuple) -> Optional[Dict[str, Any]]:
        """解析Excel行数据"""
        try:
            word_data = {}
            
            # 映射列名到字段
            field_mapping = {
                '字序号': 'word_id',
                '拉丁写法': 'latin_form',
                '音标': 'phonetic',
                '词性': 'word_type',
                '释义': 'definitions',
                '例句': 'examples',
                '备注': 'notes'
            }
            
            for i, header in enumerate(headers):
                if header in field_mapping and i < len(row):
                    value = row[i]
                    if value:
                        field_name = field_mapping[header]
                        
                        if field_name == 'definitions':
                            # 解析释义（用分号分隔）
                            word_data[field_name] = [d.strip() for d in str(value).split(';') if d.strip()]
                        elif field_name == 'examples':
                            # 解析例句（用分号分隔）
                            examples = []
                            for ex in str(value).split(';'):
                                ex = ex.strip()
                                if ex:
                                    # 简单解析例句和翻译
                                    if '(' in ex and ')' in ex:
                                        example_text = ex.split('(')[0].strip()
                                        translation = ex.split('(')[1].split(')')[0].strip()
                                        examples.append({'text': example_text, 'translation': translation})
                                    else:
                                        examples.append({'text': ex, 'translation': ''})
                            word_data[field_name] = examples
                        else:
                            word_data[field_name] = str(value).strip()
            
            # 验证必需字段
            if not word_data.get('word_id') or not word_data.get('latin_form'):
                return None
            
            return word_data
            
        except Exception as e:
            self.logger.error(f"解析Excel行数据失败: {e}")
            return None
    
    def _parse_csv_row(self, row: Dict[str, str]) -> Optional[Dict[str, Any]]:
        """解析CSV行数据"""
        try:
            word_data = {}
            
            # 映射列名到字段
            field_mapping = {
                'word_id': 'word_id',
                'latin_form': 'latin_form',
                'phonetic': 'phonetic',
                'word_type': 'word_type',
                'definitions': 'definitions',
                'examples': 'examples',
                'notes': 'notes'
            }
            
            for csv_field, data_field in field_mapping.items():
                if csv_field in row and row[csv_field]:
                    value = row[csv_field].strip()
                    
                    if data_field == 'definitions':
                        word_data[data_field] = [d.strip() for d in value.split(';') if d.strip()]
                    elif data_field == 'examples':
                        examples = []
                        for ex in value.split(';'):
                            ex = ex.strip()
                            if ex:
                                if '|' in ex:
                                    parts = ex.split('|', 1)
                                    examples.append({'text': parts[0].strip(), 'translation': parts[1].strip()})
                                else:
                                    examples.append({'text': ex, 'translation': ''})
                        word_data[data_field] = examples
                    else:
                        word_data[data_field] = value
            
            # 验证必需字段
            if not word_data.get('word_id') or not word_data.get('latin_form'):
                return None
            
            return word_data
            
        except Exception as e:
            self.logger.error(f"解析CSV行数据失败: {e}")
            return None
    
    def _validate_word_data(self, word_data: Dict[str, Any]) -> bool:
        """验证词条数据"""
        required_fields = ['word_id', 'latin_form']
        
        for field in required_fields:
            if not word_data.get(field):
                return False
        
        # 验证释义格式
        if 'definitions' in word_data:
            if not isinstance(word_data['definitions'], list):
                return False
        
        # 验证例句格式
        if 'examples' in word_data:
            if not isinstance(word_data['examples'], list):
                return False
            
            for example in word_data['examples']:
                if isinstance(example, dict):
                    if not example.get('text'):
                        return False
                elif isinstance(example, str):
                    if not example.strip():
                        return False
        
        return True


# 全局导入服务实例
import_service = ImportService()
