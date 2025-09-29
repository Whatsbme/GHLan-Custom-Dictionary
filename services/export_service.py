"""
导出服务
Export Service
"""

import os
from pathlib import Path
from typing import List, Dict, Any
import logging

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from models import WordEntry
from services.database_service import db_service
from app.config import config


class ExportService:
    """导出服务类"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.export_dir = Path(config.exports_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_pdf(self, word_entries: List[WordEntry] = None, filename: str = None, progress_callback=None) -> str:
        """导出为PDF格式"""
        try:
            if word_entries is None:
                from services.dictionary_service import dictionary_service
                word_entries = dictionary_service.get_all_word_entries()
            
            if not filename:
                filename = f"dictionary_export_{self._get_timestamp()}.pdf"
            
            filepath = self.export_dir / filename
            total_entries = len(word_entries)
            
            # 创建PDF文档
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # 标题样式
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # 居中
            )
            
            # 添加标题
            story.append(Paragraph("离线词典导出", title_style))
            story.append(Spacer(1, 20))
            
            # 添加词条
            for word_entry in word_entries:
                # 词条标题
                word_title = f"{word_entry.word_id} - {word_entry.latin_form}"
                if word_entry.phonetic:
                    word_title += f" [{word_entry.phonetic}]"
                
                story.append(Paragraph(word_title, styles['Heading2']))
                
                # 基本信息表格
                info_data = [
                    ['字序号', word_entry.word_id],
                    ['拉丁写法', word_entry.latin_form],
                    ['音标', word_entry.phonetic or ''],
                    ['词性', word_entry.word_type or ''],
                ]
                
                info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
                info_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(info_table)
                story.append(Spacer(1, 12))
                
                # 释义
                if word_entry.definitions:
                    story.append(Paragraph("释义:", styles['Heading3']))
                    for i, definition in enumerate(word_entry.definitions, 1):
                        story.append(Paragraph(f"{i}. {definition.definition_text}", styles['Normal']))
                    story.append(Spacer(1, 12))
                
                # 例句
                if word_entry.examples:
                    story.append(Paragraph("例句:", styles['Heading3']))
                    for i, example in enumerate(word_entry.examples, 1):
                        story.append(Paragraph(f"{i}. {example.example_text}", styles['Normal']))
                        if example.translation:
                            story.append(Paragraph(f"   翻译: {example.translation}", styles['Normal']))
                    story.append(Spacer(1, 12))
                
                # 备注
                if word_entry.notes:
                    story.append(Paragraph("备注:", styles['Heading3']))
                    story.append(Paragraph(word_entry.notes, styles['Normal']))
                    story.append(Spacer(1, 12))
                
                story.append(Spacer(1, 20))
            
            # 构建PDF
            doc.build(story)
            
            self.logger.info(f"PDF导出成功: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"PDF导出失败: {e}")
            raise
    
    def export_to_excel(self, word_entries: List[WordEntry], filename: str = None) -> str:
        """导出为Excel格式"""
        try:
            if not filename:
                filename = f"dictionary_export_{self._get_timestamp()}.xlsx"
            
            filepath = self.export_dir / filename
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "词条列表"
            
            # 设置标题行
            headers = ['字序号', '拉丁写法', '音标', '词性', '释义', '例句', '备注', '创建时间']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 添加数据
            for row, word_entry in enumerate(word_entries, 2):
                ws.cell(row=row, column=1, value=word_entry.word_id)
                ws.cell(row=row, column=2, value=word_entry.latin_form)
                ws.cell(row=row, column=3, value=word_entry.phonetic or '')
                ws.cell(row=row, column=4, value=word_entry.word_type or '')
                
                # 合并释义
                definitions = [defn.definition_text for defn in word_entry.definitions]
                ws.cell(row=row, column=5, value='; '.join(definitions))
                
                # 合并例句
                examples = [f"{ex.example_text} ({ex.translation or ''})" for ex in word_entry.examples]
                ws.cell(row=row, column=6, value='; '.join(examples))
                
                ws.cell(row=row, column=7, value=word_entry.notes or '')
                ws.cell(row=row, column=8, value=word_entry.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # 设置列宽
            column_widths = [12, 20, 15, 10, 40, 40, 30, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # 保存文件
            wb.save(str(filepath))
            
            self.logger.info(f"Excel导出成功: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Excel导出失败: {e}")
            raise
    
    def export_to_dict_format(self, word_entries: List[WordEntry], filename: str = None) -> str:
        """导出为字典格式（纯文本）"""
        try:
            if not filename:
                filename = f"dictionary_export_{self._get_timestamp()}.txt"
            
            filepath = self.export_dir / filename
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("离线词典导出\n")
                f.write("=" * 50 + "\n\n")
                
                for word_entry in word_entries:
                    # 词条标题
                    f.write(f"{word_entry.word_id} - {word_entry.latin_form}")
                    if word_entry.phonetic:
                        f.write(f" [{word_entry.phonetic}]")
                    f.write("\n")
                    f.write("-" * 30 + "\n")
                    
                    # 基本信息
                    f.write(f"字序号: {word_entry.word_id}\n")
                    f.write(f"拉丁写法: {word_entry.latin_form}\n")
                    if word_entry.phonetic:
                        f.write(f"音标: {word_entry.phonetic}\n")
                    if word_entry.word_type:
                        f.write(f"词性: {word_entry.word_type}\n")
                    
                    # 释义
                    if word_entry.definitions:
                        f.write("\n释义:\n")
                        for i, definition in enumerate(word_entry.definitions, 1):
                            f.write(f"{i}. {definition.definition_text}\n")
                    
                    # 例句
                    if word_entry.examples:
                        f.write("\n例句:\n")
                        for i, example in enumerate(word_entry.examples, 1):
                            f.write(f"{i}. {example.example_text}\n")
                            if example.translation:
                                f.write(f"   翻译: {example.translation}\n")
                    
                    # 备注
                    if word_entry.notes:
                        f.write(f"\n备注: {word_entry.notes}\n")
                    
                    f.write("\n" + "=" * 50 + "\n\n")
            
            self.logger.info(f"字典格式导出成功: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"字典格式导出失败: {e}")
            raise
    
    def _get_timestamp(self) -> str:
        """获取时间戳字符串"""
        from datetime import datetime
        return datetime.now().strftime("%Y%m%d_%H%M%S")


# 全局导出服务实例
export_service = ExportService()
